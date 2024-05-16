from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from Nacidos.consultaN import Consulta
from Consulta_Galen import queryGalen

class RGeneral(object):

	def __init__(self):
		self.obj_consultaN=Consulta()
		self.obj_consultaGalen=queryGalen()

	def General(self,event,fechaI,fechaF):		
		
		rows=self.obj_consultaN.consulta_General(fechaI,fechaF)
		wb=Workbook()
		sheet=wb.active
		sheet.merge_cells('A1:K1')		
		sheet['A1']=f"REGISTRO DE RECIEN NACIDOS"
		sheet['A1'].alignment=Alignment(horizontal="center")
		

		sheet.merge_cells('L1:N1')
		sheet['L1']="Desde: "+str(fechaI)+" -  Hasta: "+str(fechaF) 

		sheet['A2']="DNI MADRE"
		sheet['B2']="NOMBRES Y APELLIDOS"
		sheet['C2']="GRUPO_FACTOR"
		sheet['D2']="RPM"
		sheet['E2']="HTA"		
		sheet['F2']="ITU3_TRIMESTRE"		
		sheet['G2']="DOSIS ITU"		
		sheet['H2']="CPN"		
		sheet['I2']="OBSERVACION"		
		sheet['J2']="PROCEDENCIA"
		sheet['K2']="INGRESO SO"
		sheet['L2']="EGRESO SO"
		sheet['M2']="INGRESO SP"
		sheet['N2']="TIPO PARTO"
		sheet['O2']="MOTIVO CESARIA"		
		sheet['P2']="HCL RN"		
		sheet['Q2']="NOMBRES Y APELLIDOS"		
		sheet['R2']="CNV"		
		sheet['S2']="PINZAMIENTO(HH:MM)"		
		sheet['T2']="CONTACTO PRECOZ"
		sheet['V2']="PROCEDENCIA"
		sheet['W2']="LME"
		sheet['X2']="PAPACANGURO"
		sheet['Y2']="PESO"
		sheet['Z2']="TALLA"
		sheet['AA2']="PC"		
		sheet['AB2']="PT"		
		sheet['AC2']="PA"		
		sheet['AD2']="PB"		
		sheet['AE2']="EXA FI"		
		sheet['AF2']="FUR"
		sheet['AG2']="APGAR1"
		sheet['AH2']="APGAR5"
		sheet['AI2']="APGAR10"
		sheet['AJ2']="TEMPERATURA"
		sheet['AK2']="PROF OCULAR"
		sheet['AL2']="VIT_K"		
		sheet['AM2']="CLASF. NUTRICIONAL"		
		sheet['AN2']="L. AMNIOTICO"		
		sheet['AO2']="KRISTELLER"		
		sheet['AP2']="MECONIO"		
		sheet['AQ2']="ORINA"
		sheet['AR2']="ASFIXIA"
		sheet['AS2']="DESTINO RN"
		sheet['AT2']="OBSERVACION RN"
		sheet['AV2']="GRUPO FACTOR"		
		sheet['AW2']="H EGRESO AIR"		
		sheet['AX2']="Fecha Nacimiento"		
		sheet['AY2']="INTERCONSULTA"
		sheet['BZ2']="RES. INTERCONSULTA"		
		sheet['BA2']="RES. ENFERMERA"		
		sheet['BB2']="RES. TEC ENFERMERA"		
		sheet['BC2']="RES. MEDICO"
		sheet['BD2']="RES.OBSTETRA"	
		
		
		nro=3
		for valor in rows:
			dnipaciente=valor.DNI					
			sheet['A'+str(nro)]=dnipaciente
			#datos paciente
			rowspaciente=self.obj_consultaGalen.query_DatosPaciente(dnipaciente)
			paciente=rowspaciente[0].PrimerNombre+" "+rowspaciente[0].ApellidoPaterno+" "+rowspaciente[0].ApellidoMaterno
			sheet['B'+str(nro)]=paciente
			sheet['C'+str(nro)]=valor.GRUPO_FACTOR
			sheet['D'+str(nro)]=valor.RPM
			sheet['E'+str(nro)]=valor.HTA
			sheet['F'+str(nro)]=valor.ITU3_TRIMESTRE
			sheet['G'+str(nro)]=valor.DOSIS_ITU
			sheet['H'+str(nro)]=valor.CPN
			sheet['I'+str(nro)]=valor.OBSERVACION
			sheet['J'+str(nro)]=valor.PROCEDENCIA
			sheet['K'+str(nro)]=valor.H_INGRESO_SOP
			sheet['L'+str(nro)]=valor.H_EGRESO_SOP
			sheet['M'+str(nro)]=valor.H_INGRESO_SALAP
			sheet['N'+str(nro)]=valor.tipo_Parto
			sheet['O'+str(nro)]=valor.MOTIVOCESARIA
			hcl=valor.HCL
			rowsRN=self.obj_consultaGalen.query_PacienteXHCL(hcl)
			rndatos=rowsRN[0].PrimerNombre+" "+rowsRN[0].ApellidoPaterno+" "+rowsRN[0].ApellidoMaterno
			sheet['P'+str(nro)]=hcl
			sheet['Q'+str(nro)]=rndatos
			sheet['R'+str(nro)]=valor.CNV
			sheet['S'+str(nro)]=valor.T_PINZA
			sheet['T'+str(nro)]=valor.CONTAC_PRECOZ
			sheet['W'+str(nro)]=valor.LME
			sheet['X'+str(nro)]=valor.CONTAC_PAPACANGURO
			sheet['Y'+str(nro)]=valor.PESO
			sheet['Z'+str(nro)]=valor.TALLA
			sheet['AA'+str(nro)]=valor.PC
			sheet['AB'+str(nro)]=valor.PT
			sheet['AC'+str(nro)]=valor.PA
			sheet['AD'+str(nro)]=valor.PB
			sheet['AE'+str(nro)]=valor.EX_FI
			sheet['AF'+str(nro)]=valor.FUR
			sheet['AG'+str(nro)]=valor.APGAR_1
			sheet['AH'+str(nro)]=valor.APGAR_5
			sheet['AI'+str(nro)]=valor.APGAR_10
			sheet['AJ'+str(nro)]=valor.TEMPERATURA
			sheet['AK'+str(nro)]=valor.PROF_OCULAR

			sheet['AL'+str(nro)]=valor.VIT_K
			sheet['AM'+str(nro)]=valor.CLASF_NUTRICIONAL
			sheet['AN'+str(nro)]=valor.L_AMNIOTICO
			sheet['AO'+str(nro)]=valor.KRISTELLER
			sheet['AP'+str(nro)]=valor.MECONIO
			sheet['AQ'+str(nro)]=valor.ORINA
			sheet['AR'+str(nro)]=valor.ASFIXIA
			sheet['AS'+str(nro)]=valor.DESTINO_RN
			sheet['AT'+str(nro)]=valor.OBS_RN
			sheet['AV'+str(nro)]=valor.GRUPORN
			sheet['AW'+str(nro)]=valor.H_EGRESO_AIRN
			sheet['AX'+str(nro)]=valor.Fecha_Nacimiento
			sheet['AY'+str(nro)]=valor.INTERCONSULTA
			sheet['AZ'+str(nro)]=valor.RESP_MEDICO_INTERCONSULTA

			dnienfermera=valor.ENFERMERA
			rowsenfermera=self.obj_consultaGalen.query_EmpleadoDNI(dnienfermera)

			if rowsenfermera:
				datosenfermera=rowsenfermera[0].Nombres+" "+rowsenfermera[0].ApellidoPaterno+" "+rowsenfermera[0].ApellidoMaterno
			else:
				datosenfermera=""

			sheet['BA'+str(nro)]=datosenfermera
			
			sheet['BB'+str(nro)]=valor.TEC_ENFERMERA
			dnienmedico=valor.MEDICO			
			rowsmedico=self.obj_consultaGalen.query_EmpleadoDNI(dnienmedico)
			if rowsmedico:
				datosmedico=rowsmedico[0].Nombres+" "+rowsmedico[0].ApellidoPaterno+" "+rowsmedico[0].ApellidoMaterno
			else:
				datosmedico=""			
			
			sheet['BC'+str(nro)]=datosmedico

			dnienobstetra=valor.OBSTETRA
			rowsobstetra=self.obj_consultaGalen.query_EmpleadoDNI(dnienobstetra)
			if rowsobstetra:
				datosobstetra=rowsobstetra[0].Nombres+" "+rowsobstetra[0].ApellidoPaterno+" "+rowsobstetra[0].ApellidoMaterno
			else:
				datosobstetra=""
							
			sheet['BD'+str(nro)]=datosobstetra
			
			nro=nro+1			

		wb.save("general.xlsx")

