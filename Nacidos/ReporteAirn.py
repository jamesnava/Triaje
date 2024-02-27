from tkcalendar import DateEntry
from tkinter import *
from tkinter import messagebox
from tkinter import ttk,filedialog
from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from Nacidos.consultaN import Consulta
from Consulta_Galen import queryGalen
class Reporte():
	def __init__(self):
		double_border_side=Side(border_style="thin")
		self.borde_caja=Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
	def TopReporte(self):
		self.Top_Reporte=Toplevel()
		self.Top_Reporte.geometry("400x150")
		self.Top_Reporte.title("Reporte AIRN")
		self.Top_Reporte.resizable(0,0)
		self.Top_Reporte.grab_set()

		label=Label(self.Top_Reporte,text="Desde: ",font=('Arial',10,'bold'))
		label.grid(row=1,column=1,pady=10)	
		self.fechaDesde=DateEntry(self.Top_Reporte,selectmode='day',date_pattern='yyyy-MM-dd')
		self.fechaDesde.grid(row=1,column=2,pady=10,padx=10)

		label=Label(self.Top_Reporte,text="Desde: ",font=('Arial',10,'bold'))
		label.grid(row=1,column=3,pady=10)	
		self.fechaHasta=DateEntry(self.Top_Reporte,selectmode='day',date_pattern='yyyy-MM-dd')
		self.fechaHasta.grid(row=1,column=4,pady=10,padx=10)

		buttonGENERAR=ttk.Button(self.Top_Reporte,text="Generar")
		buttonGENERAR.grid(row=2,column=3,pady=10)
		buttonGENERAR['command']=self.evento_Button

	def evento_Button(self):
		file_Address=filedialog.asksaveasfile(mode="w",defaultextension=".xlsx")
		self.ReporteNacidos(file_Address)
		self.Top_Reporte.destroy()
		messagebox.showinfo("Notificación","Se generó correctamente!!")

	def ReporteNacidos(self,address):
		obj_consulta=Consulta()
		obj_ConsultaGalen=queryGalen()
		fechaI=self.fechaDesde.get_date()
		fechaF=self.fechaHasta.get_date()
		rows=obj_consulta.Report_General(fechaI,fechaF)
		wb=Workbook()
		sheet=wb.active
		sheet.merge_cells('A1:K1')		
		sheet['A1']=f"REGISTRO DE SEGUIMIENTO DE RECIEN NACIDOS CON ALTA HOSPITALARIA - {str(fechaF).split('-')[0]}"
		sheet['A1'].alignment=Alignment(horizontal="center")
		sheet['A1'].border=self.borde_caja

		sheet.merge_cells('L1:N1')
		sheet['L1']="Desde: "+str(fechaI)+" -  Hasta: "+str(fechaF) 

		sheet['A2']="Nro"
		sheet['B2']="HCL"
		sheet['C2']="COD CNV"
		sheet['D2']="EE.SS DE NACIMIENTO"
		sheet.column_dimensions['E'].width = 20
		sheet['E2']="EE.SS DE ORIGEN"
		sheet.column_dimensions['F'].width = 30
		sheet['F2']="NOMBRES Y APELLIDOS DE RN"
		sheet.column_dimensions['G'].width = 15
		sheet['G2']="FECHA DE NACIMIENTO"
		sheet.column_dimensions['H'].width = 7
		sheet['H2']="HOSPITALIZADO"
		sheet.column_dimensions['I'].width = 10
		sheet['I2']="FECHA DE ALTA"

		sheet['J2']="RN CON TAMIZAJE NEONATAL"
		sheet.column_dimensions['K'].width = 30
		sheet['K2']="NOMBRES Y APELLIDOS DE LA MADRE"
		sheet['L2']="Nro DE CELULAR FAMILIARES"
		sheet.column_dimensions['M'].width = 30
		sheet['M2']="RESPONSABLES DEL ALTA RN"
		sheet.column_dimensions['N'].width = 30
		sheet['N2']="OBSERVACION"
		n=3
		nro=1
		for valor in rows:
			hclN=valor.HCL
			dniMadre=valor.DNI
			dniMedicoResponsable=valor.DR_RESPONSABLE			
			rowsGalenNino=obj_ConsultaGalen.query_PacienteXHCL(hclN)			
			rows_Madre=obj_ConsultaGalen.query_DatosPaciente(dniMadre)
			#DATOS MEDICO
			rows_Medico=obj_ConsultaGalen.query_EmpleadoDNI(dniMedicoResponsable)
					
			sheet['A'+str(n)]=nro
			sheet['B'+str(n)]=hclN
			sheet['C'+str(n)]=valor.CNV
			sheet['D'+str(n)]=""
			sheet['E'+str(n)]=valor.PROCEDENCIA
			sheet['F'+str(n)]=rowsGalenNino[0].PrimerNombre+" "+rowsGalenNino[0].ApellidoPaterno+" "+rowsGalenNino[0].ApellidoMaterno
			sheet['G'+str(n)]=rowsGalenNino[0].FechaNacimiento
			sheet['H'+str(n)]=""
			sheet['I'+str(n)]=valor.FECHA_ALTA
			sheet['J'+str(n)]=""
			sheet['K'+str(n)]=rows_Madre[0].PrimerNombre+" "+rows_Madre[0].ApellidoPaterno+" "+rows_Madre[0].ApellidoMaterno
			sheet['L'+str(n)]=rows_Madre[0].Telefono
			sheet['M'+str(n)]=rows_Medico[0].Nombres+" "+rows_Medico[0].ApellidoPaterno+" "+rows_Medico[0].ApellidoMaterno
			sheet['N'+str(n)]=valor.OBS_RN
			nro=nro+1
			n=n+1

		wb.save(f"{address.name}")